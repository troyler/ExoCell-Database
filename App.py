import sys
import re
from distutils.filelist import FileList
from operator import itemgetter
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidgetItem
from PyQt5 import QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import xlsxwriter
import os


xlsxFiles = []
fileDictionary = {}
currentDirectoryList =[]
fileList = []
namingConv = {}
summaryList = ["Files graphed in this sheet", "    "]


class AnotherWindow(QWidget):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """
    def __init__(self, locationPath):
        super().__init__()

        self.btn = QtWidgets.QPushButton('Click to enter cell name', self)
        self.btn.move(20, 20)
        self.btn.clicked.connect(self.gettingSummaryName)
        self.le = QtWidgets.QLineEdit(self)
        self.le.move(130, 22)
        self.condCount = 6
        self.condLetterCount = 2
        self.scCount = 6
        self.scLetterCount = 2
        self.selectMsg = QtWidgets.QMessageBox()
        self.setWindowTitle("Graphing XLSX Files")





        self.setGeometry(800,200,700,400)
        self.locationPath = locationPath
        self.layout = QVBoxLayout()
        self.xlsxListWidget = QtWidgets.QListWidget()
        self.xlsxListWidget.setAlternatingRowColors(True)
        self.finalButton = QtWidgets.QPushButton("Create Summary File")
        self.finalButton.clicked.connect(lambda: self.extractingGraphs())
        self.xlsxButton = QtWidgets.QPushButton("Push to Select Files")
        self.xlsxButton.clicked.connect(lambda: self.getXLSXinTable())
        self.setLayout(self.layout)
        self.layout.addWidget(self.xlsxListWidget)
        self.layout.addWidget(self.xlsxButton)
        self.layout.addWidget(self.btn)

    def getXLSXinTable (self):
        x= 0
        try:
            saveLocation =  QtWidgets.QFileDialog.getOpenFileNames(self, "Open file", "", "Excel Files (*.xlsx)") #tuple ([list of strings], string)
            self.longPath = saveLocation[0][0][:saveLocation[0][0].rindex("/")+1]
            self.pathStep = self.longPath.split("/")
            self.relativePath = saveLocation[0][0][saveLocation[0][0].rindex("/")+1:]
        except IndexError:     
            self.selectMsg.setText("Error, must choose files")
            self.selectMsg.exec_()
            return
             

        for each in saveLocation[0]:
            if ".DS_Store" in each:
                saveLocation.remove(".DS_Store")
            if each not in xlsxFiles:    
                while x < len(saveLocation[0]):
                    xlsxFile = saveLocation[0][x]
                    xlsxFiles.append(xlsxFile)
                    shortName = xlsxFile[xlsxFile.rindex("/")+1:]
                    summaryList.append(shortName)
                    self.xlsxListWidget.addItem(shortName)
                    x+=1


    def gettingSummaryName(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'input dialog', 'Enter Name for Summary File')
        if ok:
            self.sumName = text
            print(text)
            self.layout.addWidget(self.finalButton)
    
    def extractingGraphs(self):
        workbook = xlsxwriter.Workbook('{}{}_Summary_File.xlsx'.format(self.longPath,self.sumName))
        workbook.add_worksheet("File Breakdown")
        workbook.add_worksheet("Conditioning Plots")
        workbook.add_worksheet("Scan Current Plots")
        fileBreakdown = workbook.get_worksheet_by_name("File Breakdown")
        condPlots = workbook.get_worksheet_by_name("Conditioning Plots")
        scPlots = workbook.get_worksheet_by_name("Scan Current Plots")


        for file in xlsxFiles:
            openFile = openpyxl.load_workbook(file, read_only=True)
            dataSet = openFile.active
            activeSheet = dataSet.title
            relativePath = file[file.rindex("/")+1:]
            seriesRef = "[{}]{}".format(relativePath,activeSheet)
            columnList = ["E","M"]
            scCount = 0

            fileBreakdown.write_column('A1', summaryList)

            if "Cond" in file and "SC" not in file:
                chart1 = workbook.add_chart({'type': 'scatter'})
                print(seriesRef)
            
                chart1.add_series({
                    'name':       [f"{seriesRef}", 0, 2],
                    'categories': [f"{seriesRef}", 1, 0, dataSet.max_row, 0],
                    'values':     [f"{seriesRef}", 1, 2, dataSet.max_row, 2],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart1.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart1.set_legend({'position': 'none'})
                chart1.set_x_axis({'name': 'Time (Sec)', 'major_gridlines': {'visible': True}})
                chart1.set_y_axis({'name': 'I (mA/cm²)'})
                condPlots.insert_chart('{}{}'.format(columnList[self.condLetterCount%2],self.condCount), chart1)
                if self.condLetterCount%2 == 0 :
                    self.condLetterCount += 1
                else:
                    self.condCount += 17
                    self.condLetterCount += 1

                openFile.close()

            elif "SC" in file:
                chart1 = workbook.add_chart({'type': 'scatter'})
                print(seriesRef)
            
                chart1.add_series({
                    'name':       [f"{seriesRef}", 0, 2],
                    'categories': [f"{seriesRef}", 1, 2, dataSet.max_row, 2],
                    'values':     [f"{seriesRef}", 1, 5, dataSet.max_row, 5],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart1.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart1.set_legend({'position': 'none'})
                chart1.set_x_axis({'name': 'I (mA/cm²)', 'major_gridlines': {'visible': True}})
                chart1.set_y_axis({'name': 'E_Stack (V)'})
                scPlots.insert_chart('{}{}'.format(columnList[self.scLetterCount%2],self.scCount), chart1)
                if self.scLetterCount%2 == 0 :
                    self.scLetterCount += 1
                else:
                    self.scCount += 17
                    self.scLetterCount += 1

                chart2 = workbook.add_chart({'type': 'scatter'})
                print(seriesRef)
            
                chart2.add_series({
                    'name':       [f"{seriesRef}", 0, 2],
                    'categories': [f"{seriesRef}", 1, 2, dataSet.max_row, 2],
                    'values':     [f"{seriesRef}", 1, 4, dataSet.max_row, 4],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart2.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart2.set_legend({'position': 'none'})
                chart2.set_x_axis({'name': 'I (mA/cm²)', 'major_gridlines': {'visible': True}})
                chart2.set_y_axis({'name': 'Power (mW/cmÂ²)'})
                scPlots.insert_chart('{}{}'.format(columnList[self.scLetterCount%2],self.scCount), chart2)

                if self.scLetterCount%2 == 0 :
                    self.scLetterCount += 1
                else:
                    self.scCount += 17
                    self.scLetterCount += 1

                

                openFile.close()
        workbook.close()

class MainWindow(QMainWindow):
    

    def __init__(self):
        super(MainWindow, self).__init__()
        self.setGeometry(0,300,1000,700)

        self.w = None  # No external window yet.
        self.button = QtWidgets.QPushButton("Push for Window")
        self.button.clicked.connect(lambda: self.show_new_window(self.saveLocation))

        self.msg = QtWidgets.QMessageBox()

        self.chooseSaveLocationButton = QtWidgets.QPushButton("Press to choose save location")
        self.chooseSaveLocationButton.clicked.connect(lambda: self.saveLocation())
        self.saveLocationStamp = QtWidgets.QLabel("Save Location")
        self.saveLocationLabel = QtWidgets.QLabel("-------------")
        self.saveLocationLabel.setWordWrap(True)
        self.saveLocationLabel.setMaximumWidth(150)

        self.setWindowTitle("TFFC Data Automation")
        self.openFileButton = QtWidgets.QPushButton("Press to open files")
        self.openFileButton.clicked.connect(lambda : self.fileChooser())
        self.fileListWidget = QtWidgets.QListWidget()
        self.fileListWidget.setAlternatingRowColors(True)
        self.namingLabel = QtWidgets.QLabel("File Breakdown")
        self.analysisButton = QtWidgets.QPushButton("Create Excel Files")     
        self.analysisButton.clicked.connect(lambda: self.fileRunner(namingConv))   
        self.fileTableBreak= QtWidgets.QTableWidget(6,6)

        self.horizontalLayout = QHBoxLayout()
    

        self.fileOpenerLayout = QVBoxLayout()
        self.fileOpenerLayout.addWidget(self.fileListWidget)
        self.fileOpenerLayout.addWidget(self.openFileButton)

        self.fileViewerLayout = QVBoxLayout()
        self.fileViewerLayout.addWidget(self.namingLabel)
        self.fileViewerLayout.addWidget(self.fileTableBreak)
        self.fileViewerLayout.addWidget(self.analysisButton)
        self.fileViewerLayout.addWidget(self.button)

        self.fileSaveLayout = QVBoxLayout()
        self.fileSaveLayout.addWidget(self.saveLocationStamp)
        self.fileSaveLayout.addWidget(self.saveLocationLabel)
        self.fileSaveLayout.addWidget(self.chooseSaveLocationButton)

        self.horizontalLayout.addLayout(self.fileOpenerLayout)
        self.horizontalLayout.addLayout(self.fileSaveLayout)
        self.horizontalLayout.addLayout(self.fileViewerLayout)

        widget = QWidget()
        widget.setLayout(self.horizontalLayout)
        self.setCentralWidget(widget)
        self.popups = []
        
       # self.fileViewerFunc()

    def show_new_window(self, xlsxLocation):
        if self.w is None:
            print(self.w)
            self.w = AnotherWindow(xlsxLocation)
            self.w.show()

        else:
            self.w.close()  # Close window.
            self.w = None  # Discard reference.

    def use_regex(self,input_text):
        pattern = re.compile(r"\d_[A-Za-z0-9]+_[A-Za-z0-9]+_[0-9]*\.[0-9]+[a-zA-Z]+_([A-Za-z0-9]+( [A-Za-z0-9]+)+)_([+-]?(?=\.\d|\d)(?:\d+)?(?:\.?\d*))(?:[eE]([+-]?\d+))?(\s([A-Za-z0-9]+\s)+)[A-Za-z0-9]+_([0-9]+(-[0-9]+)+)", re.IGNORECASE)
        print(pattern.match(input_text))
        return pattern.match(input_text)


    
#function to bring in file paths as strings in a list
    def fileChooser(self):   
        fname = QtWidgets.QFileDialog.getOpenFileNames(self, "Open file", "", "FCD Files (*.fcd)") #tuple ([list of strings], string)
        for file in fname[0]:
            self.fileTitle = file[file.rindex('/')+1:-4]   #for each file in self.fname[0], where the list of file paths is held
            if file not in fileList:   #checking to make sure current file path is not in list to hold filepaths
                self.use_regex(self.fileTitle)
                temp_file_path = " "
                temp_name = " "
                temp_file_type = ""
                temp_test_number = ""
                temp_date = ""
                temp_other = ""
                try:
                    temp_file_path = self.fileTitle.split("_")
                    if temp_file_path[0].isdigit() and len(temp_file_path[0]) <= 2 and temp_file_path[2].isdigit() == False:
                        self.temp_test_number = self.fileData[0]
                        temp_name = temp_file_path[1]
                        temp_date = temp_file_path[-1]
                        temp_other = "".join(self.fileData[3:-1])
                        testTypeCount = self.fileData[2]
                        print(testTypeCount)
                        if len(temp_name) == 8:
                            fileList.append(file)
                            print(fileList)
                            currentDirectoryList.append((file, temp_test_number, temp_file_path))
                            if "SC" or "Cond" or "OCV" in testTypeCount and len(testTypeCount) <=6:
                                namingConv[self.fileTitle] = {"Cell ID" : temp_name,
                                                                "Test Iteration" : testTypeCount,
                                                                "Other Info" : otherInfo,
                                                                "File Location" : file,
                                                                "File Title": temp_file_path,
                                                                "Cell Test Number" : temp_test_number,
                                                                "Test Date" : temp_date}
                                print(temp_name, testTypeCount, self.temp_test_number, otherInfo)
                                self.fileListWidget.addItem(self.fileTitle)
                                #print(namingConv)
                                currentDirectoryList.sort(key=itemgetter(1))
                                #print(currentDirectoryList)
                                self.fileViewerFunc()

                            else:
                                self.msg.setText(f"{self.fileTitle} was not an OCV, SC, or Cond file")
                                self.msg.exec_()  
                        else:
                            self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                            self.msg.exec_() 
                        

                    else:
                        print(self.fileTitle)
                        fileSplit = self.fileTitle.split("_")
                        number = fileSplit[-1]
                        print(number)
                        self.dateData = fileSplit[0:3]
                        temp_date = "/".join(self.dateData)
                        print(fileSplit)   
                        if len(fileSplit[3]) >= 6 and len(fileSplit[3]) <= 10:
                            fileList.append(file)
                            currentDirectoryList.append((file, number, self.fileTitle))
                            temp_name = fileSplit[3]
                            testTypeCount = fileSplit[4].replace(" ", "")
                            print(testTypeCount)
                            if "SC" or "Cond" or "OCV" in testTypeCount and len(testTypeCount) <=6:
                                temp_test_number = fileSplit[-1]
                                otherInfo = " ".join(fileSplit[4:-1])
                                namingConv[self.fileTitle] = {"Cell ID" : temp_name,
                                                                "Test Iteration" : testTypeCount,
                                                                "Other Info" : otherInfo,
                                                                "File Location" : file,
                                                                "File Title": self.fileTitle,
                                                                "Cell Test Number" : temp_test_number,
                                                                "Test Date" : temp_date}
                                print(temp_name, testTypeCount, temp_test_number, otherInfo)
                                self.fileListWidget.addItem(self.fileTitle)
                                #print(namingConv)
                                currentDirectoryList.sort(key=itemgetter(1))
                                #print(currentDirectoryList)
                                self.fileViewerFunc()

                            else:
                                self.msg.setText(f"{self.fileTitle} was not an OCV, SC, or Cond file")
                                self.msg.exec_()  
                        else:
                            self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                            self.msg.exec_() 
                        
                except IndexError:     
                    self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                    self.msg.exec_() 
                    continue
            else: 
               self.msg.setText(f"{self.fileTitle} already selected")
               self.msg.exec_()
               continue 
               # return



    def fileViewerFunc(self):
        x=0
        y= 0
        self.keys = sorted((namingConv.keys()))

        self.horizontalHeaders = ["Test Number" , "Test Type/Iteration", "Cell ID", "Test Date", "File Title", "Other Info"]
        self.fileTableBreak.setColumnCount(len(self.horizontalHeaders))
        self.fileTableBreak.setRowCount(len(currentDirectoryList))
        while x < len(currentDirectoryList):
            try1 = currentDirectoryList.sort(key = itemgetter(1))
            fileKey = currentDirectoryList[x][2]
            self.fileTableBreak.setHorizontalHeaderLabels(self.horizontalHeaders)
            self.fileTableBreak.setItem(x,0,QTableWidgetItem(namingConv[fileKey]["Cell Test Number"]))
            self.fileTableBreak.setItem(x,1,QTableWidgetItem(namingConv[fileKey]["Test Iteration"]))
            self.fileTableBreak.setItem(x,2,QTableWidgetItem(namingConv[fileKey]["Cell ID"]))
            self.fileTableBreak.setItem(x,3,QTableWidgetItem(namingConv[fileKey]["Test Date"]))
            self.fileTableBreak.setItem(x,4,QTableWidgetItem(namingConv[fileKey]["File Title"]))
            self.fileTableBreak.setItem(x,5,QTableWidgetItem(namingConv[fileKey]["Other Info"]))
            x+=1
            y+=1
            
    def fileRunner(self, x):
        self.show_new_window(self.locationPath)
        y = 0
        while y < len(currentDirectoryList):
            self.fileAnalyzer(x[self.keys[y]]["File Location"], x[self.keys[y]]["File Title"])
            y+= 1
        #print(fileDictionary)
        self.writingToExcel()
        self.fileViewerLayout.addWidget(self.button)
        

    def fileAnalyzer(self, incoming, name):
        with open(incoming, "r", encoding='latin_1') as workingFile: 
            cleanData = []  #opening and reading the incoming file as workingFile
            for line in workingFile:              #for each line in this file 
                dataInfo = line.strip().split('\t')     #strip this line of unneeded info and split it by tab. Returns a list of each item split
           #  # # # print(dataInfo)
                if len(dataInfo) > 10 and dataInfo[0]=="Time (Sec)":  #if the len of the list is greather than ten
                    cleanData.append(dataInfo)
                elif len(dataInfo) > 10:
                    dataInfo = list(map(float,dataInfo))
                    cleanData.append(dataInfo) 
            if cleanData[0][0] == "Time (Sec)" : 
                    fileDictionary[name] = pd.DataFrame(cleanData[1:], columns=cleanData[0])
                    self.occurence = fileDictionary[name]
                    #print(fileDictionary[name])
                    self.testTitle = name
                    self.timeTitle = self.occurence.columns[0]
                    self.voltageTitle = self.occurence.columns[5]
                    self.currentDensityTitle = self.occurence.columns[2]
                    self.powerDensityTitle = self.occurence.columns[4]

    def saveLocation(self):
        self.savingLocation = QtWidgets.QFileDialog.getExistingDirectory(self, "Open file") #tuple ([list of strings], string)
        self.locationPath = self.savingLocation
        self.saveLocationLabel.setText(self.locationPath)



    def writingToExcel(self):
        x= 0
        while x < len(currentDirectoryList):
            currentFileName = namingConv[self.keys[x]]["File Title"]
            with pd.ExcelWriter(f"{self.locationPath}/{currentFileName}.xlsx") as writer:
                writer.if_sheet_exists = 'replace'
                fileDictionary[currentFileName].to_excel(writer, sheet_name = f"{currentFileName[0:25]}", engine="xlsxwriter", index = False)
                self.saveLocationStamp.setText(f"Saving {currentFileName} to")
                writer.save()
                print("saving")
                x+=1
        self.saveLocationStamp.setText("Saving Complete")


    
if  __name__ == "__main__":
   
    app = QApplication(sys.argv)  
    window = MainWindow()
    window.show()

    sys.exit(app.exec_())