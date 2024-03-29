import sys
from fileInformer import fileInfo, sc_tests, OCV_tests, cond_tests
import re
from distutils.filelist import FileList
from operator import itemgetter
from fileChooser import file_chooser
from fileInformer import OCV_tests,sc_tests,cond_tests,fileInfo
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidgetItem
from PyQt5 import QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import xlsxwriter
import os
from xlsxUtil import get_XLSX_in_table


summaryList = ["Files graphed in this sheet", "    "]
xlsxFiles = []
temp_key_criteria = ["Surface Area",
                "Hydrogen Flow",
                 "Initial Current Density", 
                 "Startup OCV",
                  "Steady State Current",
                  "Steady State Current Density",
                   "Max Power Density",
                   "Voltage at Max Power Density",
                   "Time at Max Power Density"]


temp_key_labels = ["Surface Area (cm²)",
                "Hydrogen Flow (mL/min)",
                 "Initial Current Density (mA/cmÂ²)", 
                 "Startup OCV (V)",
                  "Steady State Current (mA)",
                  "Steady State Current Density (mA/cmÂ²)",
                   "Max Power Density (mW/cmÂ²)", 
                   "Voltage at Max Power Density (V)",
                   "Time at Max Power Density (s)"]

class AnotherWindow(QWidget):
    

    def __init__(self):
        super().__init__()


        self.btn = QtWidgets.QPushButton('Click to enter cell name', self)
        self.btn.move(20, 20)
        self.btn.clicked.connect(self.gettingSummaryName)
        self.line_edit= QtWidgets.QLineEdit(self)
        self.line_edit.move(130, 22)
        self.condCount = 6
        self.condLetterCount = 2
        self.scCount = 6
        self.scLetterCount = 2
        self.longPath = ""
        self.selectMsg = QtWidgets.QMessageBox()
        self.setWindowTitle("Graphing XLSX Files")


        self.setGeometry(800,200,700,400)
        self.layout = QVBoxLayout()
        self.xlsxListWidget = QtWidgets.QListWidget()
        self.xlsxListWidget.setAlternatingRowColors(True)
        self.finalButton = QtWidgets.QPushButton("Create Summary File")
        self.finalButton.clicked.connect(lambda: self.extractingGraphs())
        self.xlsxButton = QtWidgets.QPushButton("Push to Select Files")
        self.xlsxButton.clicked.connect(lambda: self.getXLSXinTable())
        self.setLayout(self.layout)
        self.layout.addWidget(self.xlsxListWidget)
        self.layout.addWidget(self.xlsxButton)
        self.layout.addWidget(self.btn)

    def  getXLSXinTable(self):
        x= 0
        try:
            saveLocation =  QtWidgets.QFileDialog.getOpenFileNames(self, "Open file", "", "Excel Files (*.xlsx)") #tuple ([list of strings], string)
            print("save location: " , saveLocation ,"\n")
            self.longPath = saveLocation[0][0][:saveLocation[0][0].rindex("/")+1]
            print(self.longPath)
        except IndexError:     
            self.selectMsg.setText("Error, must choose files")
            self.selectMsg.exec_()
            return
             
        for each in saveLocation[0]:
            if ".DS_Store" in each:
                saveLocation.remove(".DS_Store")
            if each not in xlsxFiles:    
                while x < len(saveLocation[0]):
                    xlsxFile = saveLocation[0][x]
                    xlsxFiles.append(xlsxFile)
                    shortName = xlsxFile[xlsxFile.rindex("/")+1:]
                    summaryList.append(shortName)
                    self.xlsxListWidget.addItem(shortName)
                    x+=1


    def gettingSummaryName(self):
        text, ok = QtWidgets.QInputDialog.getText(self, 'input dialog', 'Enter Name for Summary File')
        if ok:
            self.sumName = text
            print(text)
            self.layout.addWidget(self.finalButton)
    
    def extractingGraphs(self):
        workbook = xlsxwriter.Workbook('{}{}_Summary_File.xlsx'.format(self.longPath,self.sumName))
        workbook.add_worksheet("File Breakdown")
        workbook.add_worksheet("Conditioning Plots")
        workbook.add_worksheet("Scan Current Plots")
        fileBreakdown = workbook.get_worksheet_by_name("File Breakdown")
        condPlots = workbook.get_worksheet_by_name("Conditioning Plots")
        scPlots = workbook.get_worksheet_by_name("Scan Current Plots")
        fileBreakdown.write_column('A1', summaryList)


        for file in xlsxFiles:
            openFile = openpyxl.load_workbook(file, read_only=True)
            dataSet = openFile.active
            activeSheet = dataSet.title
            relativePath = file[file.rindex("/")+1:]
            seriesRef = "\'[{}]{}\'".format(relativePath,activeSheet)
            columnList = ["E","M"]
            scCount = 0


            if "Cond" in file and "SC" not in file:
                chart1 = workbook.add_chart({'type': 'scatter'})
              #  print(seriesRef)
            
                chart1.add_series({
                    'name':       [f"{seriesRef}", 1, 0],
                    'categories': [f"{seriesRef}", 4, 0, dataSet.max_row, 0],
                    'values':     [f"{seriesRef}", 4, 2, dataSet.max_row, 2],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart1.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart1.set_legend({'layout': {
        'x':      0.1,
        'y':      0.14,
        "width" : 0.8,
        "height" : 0.1

    }})
             
                chart1.set_plotarea({
    'layout': {
        'x':      0.13,
        'y':      0.24,
        'width':  0.8,
        'height': 0.60,
    }})
                chart1.set_x_axis({'name': 'Time (Sec)', 'major_gridlines': {'visible': True}})
                chart1.set_y_axis({'name': 'I (mA/cm²)'})
                condPlots.insert_chart('{}{}'.format(columnList[self.condLetterCount%2],self.condCount), chart1)
                if self.condLetterCount%2 == 0 :
                    self.condLetterCount += 1
                else:
                    self.condCount += 17
                    self.condLetterCount += 1

                openFile.close()

            elif "SC" in file:
                chart1 = workbook.add_chart({'type': 'scatter'})
             #   print(seriesRef)
            
                chart1.add_series({
                    'name':       [f"{seriesRef}", 1, 0],
                    'categories': [f"{seriesRef}", 4, 2, dataSet.max_row, 2],
                    'values':     [f"{seriesRef}", 4, 5, dataSet.max_row, 5],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart1.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart1.set_legend({'layout': {
        'x':      0.1,
        'y':      0.14,
        "width" : 0.8,
        "height" : 0.1

    }})
             
                chart1.set_x_axis({'name': 'I (mA/cm²)', 'major_gridlines': {'visible': True}})
                chart1.set_y_axis({'name': 'E_Stack (V)'})
                chart1.set_plotarea({
    'layout': {
        'x':      0.13,
        'y':      0.24,
        'width':  0.8,
        'height': 0.60,
    }})
                scPlots.insert_chart('{}{}'.format(columnList[self.scLetterCount%2],self.scCount), chart1)
                if self.scLetterCount%2 == 0 :
                    self.scLetterCount += 1
                else:
                    self.scCount += 17
                    self.scLetterCount += 1

                chart2 = workbook.add_chart({'type': 'scatter'})
              #  print(seriesRef)
            
                chart2.add_series({
                    'name':       [f"{seriesRef}", 1, 0],
                    'categories': [f"{seriesRef}", 4, 2, dataSet.max_row, 2],
                    'values':     [f"{seriesRef}", 4, 4, dataSet.max_row, 4],
                    'marker':     {'type': 'diamond', 'size': 4},
                }) 
                chart2.set_title ({'name': '{}'.format(relativePath[0:-5]), 'name_font': {'name':'Arial', 'size':10, 'bold':True},})
                chart2.set_legend({'layout': {
        'x':      0.1,
        'y':      0.14,
        "width" : 0.8,
        "height" : 0.1

    }})
                chart2.set_x_axis({'name': 'I (mA/cm²)', 'major_gridlines': {'visible': True}})
                chart2.set_y_axis({'name': 'Power (mW/cmÂ²)'})
                chart2.set_plotarea({
    'layout': {
        'x':      0.13,
        'y':      0.24,
        'width':  0.8,
        'height': 0.60,
    }})
                scPlots.insert_chart('{}{}'.format(columnList[self.scLetterCount%2],self.scCount), chart2)

                if self.scLetterCount%2 == 0 :
                    self.scLetterCount += 1
                else:
                    self.scCount += 17
                    self.scLetterCount += 1

                    

                

                openFile.close()
        infoListValues = list(self.info.values())
        infoList = list(self.info)

        for i in range(len(self.info)):
            fileBreakdown.set_column(i+1,i+1,max(len(infoListValues[i]), len(infoList[i])))
        fileBreakdown.write_row((len(summaryList) + 5), 1, data = list(self.info))
        fileBreakdown.write_row((len(summaryList) + 6), 1, data = list(self.info.values()))
        fileBreakdown.write_row(0, 6, data = temp_key_labels)
        counter = 0
        while counter < len(temp_key_criteria):
            fileBreakdown.set_column(2,counter+6, max(len(self.surface_area[temp_key_criteria[counter]]),len(temp_key_criteria[counter])))
            fileBreakdown.write_column(2,counter+6, self.surface_area[temp_key_criteria[counter]])
            counter+=1
            
        workbook.close()   
        workbook.close()
