from ast import While
from ctypes import alignment
from distutils.filelist import FileList
from operator import itemgetter
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidgetItem
from PyQt5 import QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart, Reference, Series)
import xlsxwriter
import os

from openpyxl.drawing.image import Image

finalList =[]
cellNames = []
fileDictionary = {}
currentDirectoryList =[]
fileList = []
namingConv = {}
ocvPlots = pd.DataFrame()
scPlots = pd.DataFrame()
condPlots = pd.DataFrame()
summaryFiles = pd.DataFrame()


class AnotherWindow(QWidget):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """
    def __init__(self, locationPath):
        super().__init__()
        self.setGeometry(0,0,700,400)
        layout = QVBoxLayout()
        self.xlsxListWidget = QtWidgets.QListWidget()
        self.xlsxListWidget.setAlternatingRowColors(True)
        self.finalButton = QtWidgets.QPushButton("Analyze files")
        self.finalButton.clicked.connect(lambda: self.obtainSummaryFile())
        self.xlsxButton = QtWidgets.QPushButton("Push to Select Files")
        self.xlsxButton.clicked.connect(lambda: self.getXLSXinTable())
        self.setLayout(layout)
        layout.addWidget(self.xlsxListWidget)
        layout.addWidget(self.xlsxButton)
        layout.addWidget(self.finalButton)
    
    def getXLSXinTable (self):
        x= 0
        self.saveLocation = QtWidgets.QFileDialog.getOpenFileNames(self, "Open file", "", "Excel Files (*.xlsx)") #tuple ([list of strings], string)
        for each in self.saveLocation[0]:
            if ".DS_Store" in each:
                self.saveLocation.remove(".DS_Store")
        while x < len(self.saveLocation[0]):
            xlsxFile = self.saveLocation[0][x]
            xlsxFile = xlsxFile[xlsxFile.rindex("/")+1:]
            finalList.append(xlsxFile)
            self.xlsxListWidget.addItem(xlsxFile)
            x+=1
        print(finalList)


    def obtainSummaryFile(self):
        filePath = self.saveLocation[0][0]
        summaryFilePath = filePath[0 : filePath.rindex("/")+1]
        fileTitle = filePath[filePath.rindex("/") +1:].split("_")
        for each in fileTitle:  
            print(each)
            if "-" in each and each[-1].isdigit():
                self.summaryFile = f"{summaryFilePath}{each}_Summary File.xlsx"
                print(self.summaryFile)
                self.wb = openpyxl.Workbook(self.summaryFile)
                for file in filePath:
                    if "Cond" in file:
                        active_sheet = fileTitle[0:25]
                        self.openingXlsx(file, active_sheet)
                        self.wb.add_chart(self.openingXlsx(file, active_sheet), "A10")
                self.wb.save(self.summaryFile)


    def conditioningIteration(self, file, active):
        condGraphs = self.wb['Conditioning Plots']
        interest = "='[{}]{}'".format(file,active)
        chartObj = openpyxl.chart.ScatterChart
        xvalues = Reference(interest, (1, 2), (10, 2))
        values = Reference(interest, (1, 1), (10, 1))
        series = Series(values, xvalues=xvalues)
        chartObj.append(series)
        self.wb.add_chart(chartObj)
        self.wb.save(self.summaryFile)

    def openingXlsx(self,file, active):  
        interest = "='[{}]{}'".format(file, active)
        print(interest)
        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.style = 13
        chart.x_axis.title = 'Size'
        chart.y_axis.title = 'Percentage'

        xvalues = Reference(interest, min_col=1, min_row=2, max_row= len(ws.max_column))
        values = Reference(interest, min_col=1, min_row=1, max_row= len(ws.max_column))
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
        return chart


            










class MainWindow(QMainWindow):
    

    def __init__(self):
        super(MainWindow, self).__init__()
        self.setGeometry(0,0,100,800)

        self.w = None  # No external window yet.
        self.button = QtWidgets.QPushButton("Push for Window")
        self.button.clicked.connect(self.show_new_window)

        self.msg = QtWidgets.QMessageBox()

        self.chooseSaveLocationButton = QtWidgets.QPushButton("Press to choose save location")
        self.chooseSaveLocationButton.clicked.connect(lambda: self.saveLocation())
        self.saveLocationStamp = QtWidgets.QLabel("Save Location")
        self.saveLocationLabel = QtWidgets.QLabel("-------------")
        self.saveLocationLabel.setWordWrap(True)
        self.saveLocationLabel.setMaximumWidth(150)

        self.setWindowTitle("TFFC Data Automation")
        self.openFileButton = QtWidgets.QPushButton("Press to open files")
        self.openFileButton.clicked.connect(lambda : self.fileChooser())
        self.fileListWidget = QtWidgets.QListWidget()
        self.fileListWidget.setAlternatingRowColors(True)
        self.namingLabel = QtWidgets.QLabel("File Breakdown")
        self.analysisButton = QtWidgets.QPushButton("Create Excel Files")     
        self.analysisButton.clicked.connect(lambda: self.fileRunner(namingConv))   
        self.fileTableBreak= QtWidgets.QTableWidget(6,6)

        self.horizontalLayout = QHBoxLayout()
    

        self.fileOpenerLayout = QVBoxLayout()
        self.fileOpenerLayout.addWidget(self.fileListWidget)
        self.fileOpenerLayout.addWidget(self.openFileButton)

        self.fileViewerLayout = QVBoxLayout()
        self.fileViewerLayout.addWidget(self.namingLabel)
        self.fileViewerLayout.addWidget(self.fileTableBreak)
        self.fileViewerLayout.addWidget(self.analysisButton)
        self.fileViewerLayout.addWidget(self.button)
        

        self.fileSaveLayout = QVBoxLayout()
        self.fileSaveLayout.addWidget(self.saveLocationStamp)
        self.fileSaveLayout.addWidget(self.saveLocationLabel)
        self.fileSaveLayout.addWidget(self.chooseSaveLocationButton)

        self.horizontalLayout.addLayout(self.fileOpenerLayout)
        self.horizontalLayout.addLayout(self.fileSaveLayout)
        self.horizontalLayout.addLayout(self.fileViewerLayout)

        widget = QWidget()
        widget.setLayout(self.horizontalLayout)
        self.setCentralWidget(widget)
        self.popups = []
        
       # self.fileViewerFunc()

    def show_new_window(self, xlsxLocation):
        if self.w is None:
            print(self.w)
            self.w = AnotherWindow(xlsxLocation)
            self.w.show()

        else:
            self.w.close()  # Close window.
            self.w = None  # Discard reference.


    
#function to bring in file paths as strings in a list
    def fileChooser(self):   
        fname = QtWidgets.QFileDialog.getOpenFileNames(self, "Open file", "", "FCD Files (*.fcd)") #tuple ([list of strings], string)
        for file in fname[0]:
            self.fileTitle = file[file.rindex('/')+1:-4]   #for each file in self.fname[0], where the list of file paths is held
            if file not in fileList:   #checking to make sure current file path is not in list to hold filepaths
                try:
                    self.fileType = self.fileTitle.split("_")
                    if self.fileType[0].isdigit() and len(self.fileType[0]) == 1:
                        self.fileData = self.fileType
                        print(self.fileData)
                        print(self.fileType)
                        self.cellTestNumber = self.fileData[0]
                        cellName = self.fileData[1]
                        self.testDate = self.fileData[-1]
                        otherInfo = "".join(self.fileData[3:-1])
                        testTypeCount = self.fileData[2]
                        print(testTypeCount)
                        if len(cellName) == 8:
                            fileList.append(file)
                            print(fileList)
                            currentDirectoryList.append((file, self.cellTestNumber, self.fileTitle))
                            if "SC" or "Cond" or "OCV" in testTypeCount and len(testTypeCount) <=6:
                                namingConv[self.fileTitle] = {"Cell ID" : cellName,
                                                                "Test Iteration" : testTypeCount,
                                                                "Other Info" : otherInfo,
                                                                "File Location" : file,
                                                                "File Title": self.fileTitle,
                                                                "Cell Test Number" : self.cellTestNumber,
                                                                "Test Date" : self.testDate}
                                print(cellName, testTypeCount, self.cellTestNumber, otherInfo)
                                cellNames.append(cellName)
                                self.fileListWidget.addItem(self.fileTitle)
                                #print(namingConv)
                                currentDirectoryList.sort(key=itemgetter(1))
                                #print(currentDirectoryList)
                                self.fileViewerFunc()

                            else:
                                self.msg.setText(f"{self.fileTitle} was not an OCV, SC, or Cond file")
                                self.msg.exec_()  
                        else:
                            self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                            self.msg.exec_() 
                        

                    else:
                        print(self.fileTitle)
                        number = self.fileTitle[-1]
                        print(number)
                        fileSplit = self.fileTitle.split("_")
                        self.dateData = fileSplit[0:3]
                        self.testDate = "/".join(self.dateData)
                        print(fileSplit)   
                        if len(fileSplit[3]) >= 6 and len(fileSplit[3]) <= 10:
                            fileList.append(file)
                            currentDirectoryList.append((file, number, self.fileTitle))
                            cellName = fileSplit[3]
                            testTypeCount = fileSplit[4].replace(" ", "")
                            print(testTypeCount)
                            if "SC" or "Cond" or "OCV" in testTypeCount and len(testTypeCount) <=6:
                                cellTestNumber = str(fileSplit[-1][0])
                                otherInfo = " ".join(fileSplit[4:-1])
                                namingConv[self.fileTitle] = {"Cell ID" : cellName,
                                                                "Test Iteration" : testTypeCount,
                                                                "Other Info" : otherInfo,
                                                                "File Location" : file,
                                                                "File Title": self.fileTitle,
                                                                "Cell Test Number" : cellTestNumber,
                                                                "Test Date" : self.testDate}
                                print(cellName, testTypeCount, cellTestNumber, otherInfo)
                                cellNames.append(cellName)
                                self.fileListWidget.addItem(self.fileTitle)
                                #print(namingConv)
                                currentDirectoryList.sort(key=itemgetter(1))
                                #print(currentDirectoryList)
                                self.fileViewerFunc()

                            else:
                                self.msg.setText(f"{self.fileTitle} was not an OCV, SC, or Cond file")
                                self.msg.exec_()  
                        else:
                            self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                            self.msg.exec_() 
                        
                except IndexError:     
                    self.msg.setText(f"{self.fileTitle} had improper naming convention and was removed")
                    self.msg.exec_() 
                    continue
            else: 
               self.msg.setText(f"{self.fileTitle} already selected")
               self.msg.exec_()
               continue 
               # return



    def fileViewerFunc(self):
        x=0
        y= 0
        self.keys = sorted((namingConv.keys()))

        self.horizontalHeaders = ["Test Number" , "Test Type/Iteration", "Cell ID", "Test Date", "File Title", "Other Info"]
        self.fileTableBreak.setColumnCount(len(self.horizontalHeaders))
        self.fileTableBreak.setRowCount(len(currentDirectoryList))
        while x < len(currentDirectoryList):
            fileKey = currentDirectoryList[x][2]
            self.fileTableBreak.setHorizontalHeaderLabels(self.horizontalHeaders)
            self.fileTableBreak.setItem(x,0,QTableWidgetItem(namingConv[fileKey]["Cell Test Number"]))
            self.fileTableBreak.setItem(x,1,QTableWidgetItem(namingConv[fileKey]["Test Iteration"]))
            self.fileTableBreak.setItem(x,2,QTableWidgetItem(namingConv[fileKey]["Cell ID"]))
            self.fileTableBreak.setItem(x,3,QTableWidgetItem(namingConv[fileKey]["Test Date"]))
            self.fileTableBreak.setItem(x,4,QTableWidgetItem(namingConv[fileKey]["File Title"]))
            self.fileTableBreak.setItem(x,5,QTableWidgetItem(namingConv[fileKey]["Other Info"]))
            x+=1
            y+=1
            
    def fileRunner(self, x):
        self.show_new_window(self.locationPath)
        y = 0
        while y < len(currentDirectoryList):
            self.fileAnalyzer(x[self.keys[y]]["File Location"], x[self.keys[y]]["File Title"])
            y+= 1
        #print(fileDictionary)
        self.writingToExcel()
        self.fileViewerLayout.addWidget(self.button)
        

    def fileAnalyzer(self, incoming, name):
        with open(incoming, "r", encoding='latin_1') as workingFile: 
            cleanData = []  #opening and reading the incoming file as workingFile
            for line in workingFile:              #for each line in this file 
                dataInfo = line.strip().split('\t')     #strip this line of unneeded info and split it by tab. Returns a list of each item split
           #  # # # print(dataInfo)
                if len(dataInfo) > 10 and dataInfo[0]=="Time (Sec)":  #if the len of the list is greather than ten
                    cleanData.append(dataInfo)
                elif len(dataInfo) > 10:
                    dataInfo = list(map(float,dataInfo))
                    cleanData.append(dataInfo) 
            if cleanData[0][0] == "Time (Sec)" : 
                    fileDictionary[name] = pd.DataFrame(cleanData[1:], columns=cleanData[0])
                    self.occurence = fileDictionary[name]
                    #print(fileDictionary[name])
                    self.testTitle = name
                    self.timeTitle = self.occurence.columns[0]
                    self.voltageTitle = self.occurence.columns[5]
                    self.currentDensityTitle = self.occurence.columns[2]
                    self.powerDensityTitle = self.occurence.columns[4]

    def saveLocation(self):
        self.savingLocation = QtWidgets.QFileDialog.getExistingDirectory(self, "Open file") #tuple ([list of strings], string)
        self.locationPath = self.savingLocation
        self.saveLocationLabel.setText(self.locationPath)



    def writingToExcel(self):
        x= 0
        while x < len(currentDirectoryList):
            currentFileName = namingConv[self.keys[x]]["File Title"]
            with pd.ExcelWriter(f"{self.locationPath}/{currentFileName}.xlsx") as writer:
                writer.if_sheet_exists = 'replace'
                fileDictionary[currentFileName].to_excel(writer, sheet_name = f"{currentFileName[0:25]}", engine="xlsxwriter", index = False)
                self.saveLocationStamp.setText(f"Saving {currentFileName} to")
                writer.save()
                print("saving")
                x+=1
            while x == len(currentDirectoryList):
                with pd.ExcelWriter(f'{self.locationPath}/{namingConv[self.keys[x-1]]["Cell ID"]}_Summary File.xlsx') as writer:
                    writer.if_sheet_exists = 'replace'
                    summaryFiles.to_excel(writer, sheet_name = "Summary Files", engine="xlsxwriter")
                    scPlots.to_excel(writer, sheet_name = "Scan Current Plots", engine="xlsxwriter")
                    condPlots.to_excel(writer, sheet_name = "Conditioning Plots", engine="xlsxwriter")
                    x +=1
        self.saveLocationStamp.setText("Saving Complete")


    
if  __name__ == "__main__":
   
    app = QApplication(sys.argv)  
    window = MainWindow()
    window.show()

    sys.exit(app.exec_())